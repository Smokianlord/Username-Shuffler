import csv
import os
import random
import re
import sys
import threading
import zipfile
from datetime import datetime, timezone
from posixpath import normpath
from xml.etree import ElementTree as ET
from xml.sax.saxutils import escape
import tkinter as tk

APP_NAME = "Username Shuffler"
APP_VERSION = "v2.0.0"
DATA_EXTENSIONS = (".xlsx", ".xlsm", ".csv", ".xls")

BG = "#EEF4FF"
HEADER = "#0F172A"
CARD = "#FFFFFF"
CARD_BORDER = "#D8E2F1"
SHADOW = "#C7D2FE"
TEXT = "#0F172A"
MUTED = "#64748B"
PRIMARY = "#2563EB"
PRIMARY_HOVER = "#1D4ED8"
GREEN = "#16A34A"
GREEN_HOVER = "#15803D"
CYAN = "#0891B2"
CYAN_HOVER = "#0E7490"
ORANGE = "#EA580C"
ORANGE_HOVER = "#C2410C"
SLATE = "#475569"
SLATE_HOVER = "#334155"

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def qn(namespace: str, tag: str) -> str:
    return f"{{{namespace}}}{tag}"


def resource_path(relative_path: str) -> str:
    try:
        base_path = sys._MEIPASS  # type: ignore[attr-defined]
    except Exception:
        base_path = os.path.abspath(os.path.dirname(__file__))
    return os.path.join(base_path, relative_path)


def get_app_folder() -> str:
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def clean_username(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "null"}:
        return ""
    return text


def find_data_file(folder: str) -> str | None:
    files: list[str] = []
    for file_name in os.listdir(folder):
        lower = file_name.lower()
        if file_name.startswith("~$") or file_name.startswith("."):
            continue
        if lower.endswith(DATA_EXTENSIONS):
            files.append(file_name)

    if not files:
        return None

    priority = ["usernames.xlsx", "usernames.xlsm", "usernames.csv", "username.xlsx", "username.csv"]
    for preferred in priority:
        for file_name in files:
            if file_name.lower() == preferred:
                return os.path.join(folder, file_name)

    files.sort(key=str.lower)
    return os.path.join(folder, files[0])


def read_csv_file(path: str) -> list[str]:
    usernames: list[str] = []
    encodings = ("utf-8-sig", "utf-8", "cp1252")
    last_error: Exception | None = None

    for encoding in encodings:
        try:
            with open(path, "r", encoding=encoding, newline="") as handle:
                reader = csv.reader(handle)
                for row in reader:
                    for cell in row:
                        name = clean_username(cell)
                        if name:
                            usernames.append(name)
            return usernames
        except UnicodeDecodeError as exc:
            last_error = exc

    if last_error:
        raise last_error
    return usernames


def _relationship_target(base_folder: str, target: str) -> str:
    if target.startswith("/"):
        return normpath(target.lstrip("/"))
    return normpath(f"{base_folder}/{target}")


def _first_worksheet_path(zf: zipfile.ZipFile) -> str:
    workbook_xml = ET.fromstring(zf.read("xl/workbook.xml"))
    sheets = workbook_xml.find(qn(MAIN_NS, "sheets"))
    if sheets is None or not list(sheets):
        raise RuntimeError("No worksheet found inside this Excel file.")

    first_sheet = list(sheets)[0]
    relationship_id = first_sheet.attrib.get(qn(REL_NS, "id"))
    if not relationship_id:
        return "xl/worksheets/sheet1.xml"

    rels_xml = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    for rel in rels_xml:
        if rel.attrib.get("Id") == relationship_id:
            return _relationship_target("xl", rel.attrib.get("Target", "worksheets/sheet1.xml"))

    return "xl/worksheets/sheet1.xml"


def _read_shared_strings(zf: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in zf.namelist():
        return []

    root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
    values: list[str] = []
    for item in root.findall(qn(MAIN_NS, "si")):
        pieces: list[str] = []
        for text_node in item.iter(qn(MAIN_NS, "t")):
            pieces.append(text_node.text or "")
        values.append("".join(pieces))
    return values


def read_xlsx_file(path: str) -> list[str]:
    usernames: list[str] = []

    with zipfile.ZipFile(path, "r") as zf:
        shared_strings = _read_shared_strings(zf)
        sheet_path = _first_worksheet_path(zf)
        if sheet_path not in zf.namelist():
            raise RuntimeError("Could not find the first worksheet inside this Excel file.")

        root = ET.fromstring(zf.read(sheet_path))
        sheet_data = root.find(qn(MAIN_NS, "sheetData"))
        if sheet_data is None:
            return []

        for row in sheet_data.findall(qn(MAIN_NS, "row")):
            for cell in row.findall(qn(MAIN_NS, "c")):
                cell_type = cell.attrib.get("t", "")
                value = ""

                if cell_type == "s":
                    node = cell.find(qn(MAIN_NS, "v"))
                    if node is not None and node.text is not None:
                        try:
                            value = shared_strings[int(node.text)]
                        except (ValueError, IndexError):
                            value = ""
                elif cell_type == "inlineStr":
                    pieces: list[str] = []
                    inline = cell.find(qn(MAIN_NS, "is"))
                    if inline is not None:
                        for text_node in inline.iter(qn(MAIN_NS, "t")):
                            pieces.append(text_node.text or "")
                    value = "".join(pieces)
                else:
                    node = cell.find(qn(MAIN_NS, "v"))
                    value = node.text if node is not None and node.text is not None else ""

                name = clean_username(value)
                if name:
                    usernames.append(name)

    return usernames


def read_excel_file(path: str) -> list[str]:
    lower = path.lower()
    if lower.endswith((".xlsx", ".xlsm")):
        return read_xlsx_file(path)
    if lower.endswith(".xls"):
        raise RuntimeError("Old .xls files are not supported. Save it as .xlsx or .csv and reload.")
    raise RuntimeError("Unsupported file type.")


def load_usernames_from_folder(folder: str) -> tuple[list[str], str | None, str | None]:
    path = find_data_file(folder)
    if not path:
        return [], None, "No Excel/CSV file found. Add usernames below and save to create usernames.xlsx."

    lower = path.lower()
    if lower.endswith(".csv"):
        usernames = read_csv_file(path)
    else:
        usernames = read_excel_file(path)

    if not usernames:
        return [], path, "The selected file has no usernames yet. Add usernames below and save."

    return usernames, path, None


def parse_new_usernames(raw_text: str) -> list[str]:
    parts: list[str] = []
    for line in raw_text.splitlines():
        line = line.strip()
        if not line:
            continue
        parts.extend(re.split(r"[,;]+", line))
    return [name for name in (clean_username(part) for part in parts) if name]


def write_simple_xlsx(path: str, names: list[str]) -> None:
    now = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    rows: list[str] = []
    for row_index, name in enumerate(names, start=1):
        safe = escape(name, {"\"": "&quot;"})
        rows.append(
            f'<row r="{row_index}"><c r="A{row_index}" t="inlineStr"><is><t xml:space="preserve">{safe}</t></is></c></row>'
        )
    sheet_xml = "".join(rows)
    dimension = f"A1:A{max(1, len(names))}"

    files = {
        "[Content_Types].xml": '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>
  <Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>
</Types>''',
        "_rels/.rels": '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>
  <Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>
</Relationships>''',
        "docProps/app.xml": '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes"><Application>Username Shuffler</Application></Properties>''',
        "docProps/core.xml": f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" xmlns:dcmitype="http://purl.org/dc/dcmitype/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"><dc:creator>Username Shuffler</dc:creator><cp:lastModifiedBy>Username Shuffler</cp:lastModifiedBy><dcterms:created xsi:type="dcterms:W3CDTF">{now}</dcterms:created><dcterms:modified xsi:type="dcterms:W3CDTF">{now}</dcterms:modified></cp:coreProperties>''',
        "xl/workbook.xml": '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"><sheets><sheet name="Usernames" sheetId="1" r:id="rId1"/></sheets></workbook>''',
        "xl/_rels/workbook.xml.rels": '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/><Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/></Relationships>''',
        "xl/styles.xml": '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts><fills count="1"><fill><patternFill patternType="none"/></fill></fills><borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders><cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs><cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs><cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles></styleSheet>''',
        "xl/worksheets/sheet1.xml": f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"><dimension ref="{dimension}"/><sheetViews><sheetView workbookViewId="0"/></sheetViews><sheetFormatPr defaultRowHeight="15"/><cols><col min="1" max="1" width="36" customWidth="1"/></cols><sheetData>{sheet_xml}</sheetData></worksheet>''',
    }

    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for file_name, content in files.items():
            zf.writestr(file_name, content)


def append_usernames_to_file(path: str, names: list[str], existing_names: list[str] | None = None) -> str:
    lower = path.lower()

    if lower.endswith(".csv"):
        with open(path, "a", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            for name in names:
                writer.writerow([name])
        return path

    if lower.endswith((".xlsx", ".xlsm")):
        try:
            from openpyxl import Workbook, load_workbook  # type: ignore
        except Exception:
            combined = list(existing_names or []) + names
            write_simple_xlsx(path, combined)
            return path

        if os.path.exists(path):
            workbook = load_workbook(path)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Usernames"

        for name in names:
            sheet.append([name])

        workbook.save(path)
        workbook.close()
        return path

    new_path = os.path.join(get_app_folder(), "usernames.xlsx")
    write_simple_xlsx(new_path, list(existing_names or []) + names)
    return new_path



class UsernameShufflerApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.folder = get_app_folder()
        self.usernames: list[str] = []
        self.loaded_file: str | None = None
        self.result_text = ""

        self.count_var = tk.StringVar()
        self.status_var = tk.StringVar(value="Starting...")
        self.file_var = tk.StringVar(value="Looking for sheet...")
        self.total_var = tk.StringVar(value="0")
        self.pick_var = tk.StringVar(value="0")
        self._trace_ready = False

        self.configure_window()
        self.build_ui()
        self.count_var.trace_add("write", self.on_count_changed)
        self._trace_ready = True
        self.load_data_async(initial=True)

    def configure_window(self) -> None:
        self.root.title(f"{APP_NAME} {APP_VERSION}")
        self.root.geometry("980x640")
        self.root.minsize(900, 560)
        self.root.configure(bg=BG)

        self._icon_handles = []
        self.apply_window_icon()
        # Windows/Tk sometimes paints the default title-bar icon first.
        # Re-applying after the window is ready keeps the custom icon in the title bar.
        self.root.after(50, self.apply_window_icon)
        self.root.after(300, self.apply_window_icon)
        self.root.after(1000, self.apply_window_icon)

    def apply_window_icon(self) -> None:
        """Apply the custom icon to Tk, the Windows wrapper window, and the EXE/taskbar path.

        Tk on Windows can keep showing the default feather icon if WM_SETICON is sent to
        the child window instead of the real top-level wrapper, so this sends it to every
        related HWND we can safely reach.
        """
        titlebar_ico = os.path.normpath(resource_path("titlebar.ico"))
        icon_ico = os.path.normpath(resource_path("icon.ico"))
        icon_png = os.path.normpath(resource_path("icon.png"))
        preferred_ico = titlebar_ico if os.path.exists(titlebar_ico) else icon_ico

        try:
            if os.path.exists(preferred_ico):
                self.root.iconbitmap(default=preferred_ico)
                self.root.wm_iconbitmap(default=preferred_ico)
        except Exception:
            try:
                if os.path.exists(icon_ico):
                    self.root.iconbitmap(default=icon_ico)
                    self.root.wm_iconbitmap(default=icon_ico)
            except Exception:
                pass

        try:
            if os.path.exists(icon_png):
                self.window_icon = tk.PhotoImage(file=icon_png)
                self.root.iconphoto(True, self.window_icon)
        except Exception:
            pass

        if os.name != "nt" or not os.path.exists(preferred_ico):
            return

        try:
            self.root.update_idletasks()
            import ctypes

            user32 = ctypes.windll.user32
            IMAGE_ICON = 1
            LR_LOADFROMFILE = 0x00000010
            WM_SETICON = 0x0080
            ICON_SMALL = 0
            ICON_BIG = 1
            ICON_SMALL2 = 2
            GA_ROOT = 2

            hwnd = int(self.root.winfo_id())
            candidates = []
            for handle in (
                hwnd,
                user32.GetParent(hwnd),
                user32.GetAncestor(hwnd, GA_ROOT),
            ):
                if handle and handle not in candidates:
                    candidates.append(handle)

            small_icon = user32.LoadImageW(None, preferred_ico, IMAGE_ICON, 16, 16, LR_LOADFROMFILE)
            big_icon = user32.LoadImageW(None, preferred_ico, IMAGE_ICON, 32, 32, LR_LOADFROMFILE)

            for handle in candidates:
                if small_icon:
                    user32.SendMessageW(handle, WM_SETICON, ICON_SMALL, small_icon)
                    user32.SendMessageW(handle, WM_SETICON, ICON_SMALL2, small_icon)
                if big_icon:
                    user32.SendMessageW(handle, WM_SETICON, ICON_BIG, big_icon)

            if small_icon:
                self._icon_handles.append(small_icon)
            if big_icon:
                self._icon_handles.append(big_icon)
        except Exception:
            pass

    def build_ui(self) -> None:
        self.root.grid_rowconfigure(1, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
        self.build_header()

        body = tk.Frame(self.root, bg=BG)
        body.grid(row=1, column=0, sticky="nsew", padx=16, pady=12)
        body.grid_columnconfigure(0, weight=1)
        body.grid_rowconfigure(0, weight=3)
        body.grid_rowconfigure(1, weight=2)

        # TOP AREA: result and copy are permanently visible in the default window.
        top = tk.Frame(body, bg=BG)
        top.grid(row=0, column=0, sticky="nsew")
        top.grid_columnconfigure(0, weight=1)
        top.grid_columnconfigure(1, weight=2)
        top.grid_rowconfigure(0, weight=1)
        self.build_control_card(top, 0, 0)
        self.build_result_card(top, 0, 1)

        # BOTTOM AREA: secondary tools only. Even if this area is small, copy stays visible above.
        bottom = tk.Frame(body, bg=BG)
        bottom.grid(row=1, column=0, sticky="nsew", pady=(8, 0))
        bottom.grid_columnconfigure(0, weight=1)
        bottom.grid_columnconfigure(1, weight=2)
        bottom.grid_rowconfigure(0, weight=1)
        self.build_dataset_card(bottom, 0, 0)
        self.build_add_card(bottom, 0, 1)

        self.build_status_bar()

    def build_header(self) -> None:
        header = tk.Frame(self.root, bg=HEADER, height=78)
        header.grid(row=0, column=0, sticky="ew")
        header.grid_columnconfigure(1, weight=1)
        header.grid_propagate(False)

        logo_holder = tk.Frame(header, bg=HEADER)
        logo_holder.grid(row=0, column=0, padx=(22, 14), pady=14, sticky="w")
        try:
            self.logo_image = tk.PhotoImage(file=resource_path("icon.png"))
            scale = max(1, self.logo_image.width() // 46)
            self.logo_image = self.logo_image.subsample(scale)
            tk.Label(logo_holder, image=self.logo_image, bg=HEADER).pack()
        except Exception:
            tk.Label(logo_holder, text="US", bg=PRIMARY, fg="white", font=("Segoe UI", 14, "bold"), padx=10, pady=8).pack()

        title_area = tk.Frame(header, bg=HEADER)
        title_area.grid(row=0, column=1, sticky="w")
        tk.Label(
            title_area,
            text=f"{APP_NAME} {APP_VERSION}",
            bg=HEADER,
            fg="white",
            font=("Segoe UI", 19, "bold"),
        ).pack(anchor="w")
        tk.Label(
            title_area,
            text="Type a number and copy usernames instantly.",
            bg=HEADER,
            fg="#CBD5E1",
            font=("Segoe UI", 10),
        ).pack(anchor="w", pady=(2, 0))

        tk.Label(
            header,
            text="Auto Shuffle ON",
            bg="#DCFCE7",
            fg="#166534",
            font=("Segoe UI", 10, "bold"),
            padx=12,
            pady=6,
        ).grid(row=0, column=2, padx=22, sticky="e")

    def build_status_bar(self) -> None:
        status_bar = tk.Frame(self.root, bg="#DBEAFE", height=34)
        status_bar.grid(row=2, column=0, sticky="ew")
        status_bar.grid_columnconfigure(0, weight=1)
        status_bar.grid_propagate(False)
        tk.Label(
            status_bar,
            textvariable=self.status_var,
            bg="#DBEAFE",
            fg="#1E3A8A",
            font=("Segoe UI", 10, "bold"),
            anchor="w",
            padx=18,
        ).grid(row=0, column=0, sticky="nsew", ipady=7)

    def card(self, parent: tk.Widget, row: int, column: int, *, sticky: str = "nsew") -> tk.Frame:
        shadow = tk.Frame(parent, bg=SHADOW)
        shadow.grid(row=row, column=column, sticky=sticky, padx=7, pady=7)
        shadow.grid_columnconfigure(0, weight=1)
        shadow.grid_rowconfigure(0, weight=1)
        inner = tk.Frame(shadow, bg=CARD, highlightbackground=CARD_BORDER, highlightthickness=1)
        inner.grid(row=0, column=0, sticky="nsew", padx=(0, 4), pady=(0, 4))
        return inner

    def build_control_card(self, parent: tk.Widget, row: int, column: int) -> None:
        card = self.card(parent, row, column)
        card.grid_columnconfigure(0, weight=1)
        card.grid_rowconfigure(4, weight=1)

        tk.Label(card, text="Pick Amount", bg=CARD, fg=TEXT, font=("Segoe UI", 15, "bold")).grid(
            row=0, column=0, sticky="w", padx=18, pady=(16, 4)
        )
        tk.Label(card, text="Type only the number. No Enter key needed.", bg=CARD, fg=MUTED, font=("Segoe UI", 9)).grid(
            row=1, column=0, sticky="w", padx=18, pady=(0, 10)
        )

        self.count_entry = tk.Entry(
            card,
            textvariable=self.count_var,
            justify="center",
            bg="#F8FAFC",
            fg=TEXT,
            insertbackground=TEXT,
            font=("Segoe UI", 30, "bold"),
            relief="solid",
            bd=1,
            highlightthickness=2,
            highlightbackground="#BFDBFE",
            highlightcolor=PRIMARY,
        )
        self.count_entry.grid(row=2, column=0, sticky="ew", padx=18, pady=(0, 12), ipady=8)

        action_row = tk.Frame(card, bg=CARD)
        action_row.grid(row=3, column=0, sticky="ew", padx=18, pady=(0, 10))
        action_row.grid_columnconfigure(0, weight=1)
        action_row.grid_columnconfigure(1, weight=1)
        self.button(action_row, "Re-shuffle", PRIMARY, PRIMARY_HOVER, self.shuffle_manual).grid(
            row=0, column=0, sticky="ew", padx=(0, 6), ipady=7
        )
        self.button(action_row, "Reload Sheet", GREEN, GREEN_HOVER, lambda: self.load_data_async(initial=False)).grid(
            row=0, column=1, sticky="ew", padx=(6, 0), ipady=7
        )

        quick = tk.Frame(card, bg="#F8FAFC", highlightbackground="#E2E8F0", highlightthickness=1)
        quick.grid(row=4, column=0, sticky="nsew", padx=18, pady=(0, 16))
        quick.grid_columnconfigure(1, weight=1)
        self.metric_row(quick, 0, "File", self.file_var)
        self.metric_row(quick, 1, "Loaded", self.total_var)
        self.metric_row(quick, 2, "Result", self.pick_var)

    def build_result_card(self, parent: tk.Widget, row: int, column: int) -> None:
        card = self.card(parent, row, column)
        card.grid_columnconfigure(0, weight=1)
        card.grid_rowconfigure(1, weight=1)

        header = tk.Frame(card, bg=CARD)
        header.grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 8))
        header.grid_columnconfigure(0, weight=1)
        tk.Label(header, text="Username Copy Area", bg=CARD, fg=TEXT, font=("Segoe UI", 15, "bold")).grid(row=0, column=0, sticky="w")
        self.copy_btn = self.button(header, "COPY RESULT", CYAN, CYAN_HOVER, self.copy_result, width=16)
        self.copy_btn.grid(row=0, column=1, sticky="e")

        result_frame = tk.Frame(card, bg="#D7E0EE")
        result_frame.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 16))
        result_frame.grid_columnconfigure(0, weight=1)
        result_frame.grid_rowconfigure(0, weight=1)
        self.output_box = tk.Text(
            result_frame,
            height=8,
            wrap="word",
            bg="#F8FAFC",
            fg=TEXT,
            relief="flat",
            font=("Consolas", 13),
            padx=12,
            pady=12,
            undo=False,
        )
        self.output_box.grid(row=0, column=0, sticky="nsew", padx=(0, 1), pady=(0, 1))
        self.output_box.configure(state="disabled")

    def build_dataset_card(self, parent: tk.Widget, row: int, column: int) -> None:
        card = self.card(parent, row, column)
        card.grid_columnconfigure(0, weight=1)
        card.grid_columnconfigure(1, weight=1)
        tk.Label(card, text="Dataset", bg=CARD, fg=TEXT, font=("Segoe UI", 13, "bold")).grid(
            row=0, column=0, columnspan=2, sticky="w", padx=18, pady=(15, 8)
        )
        self.metric_row(card, 1, "Current file", self.file_var)
        self.metric_row(card, 2, "Loaded usernames", self.total_var)
        self.metric_row(card, 3, "Current result", self.pick_var)


    def metric_row(self, parent: tk.Widget, row: int, label: str, value_var: tk.StringVar) -> None:
        tk.Label(parent, text=label, bg=parent.cget("bg"), fg=MUTED, font=("Segoe UI", 10)).grid(
            row=row, column=0, sticky="w", padx=12 if str(parent.cget("bg")) == "#F8FAFC" else 18, pady=5
        )
        tk.Label(parent, textvariable=value_var, bg=parent.cget("bg"), fg=TEXT, font=("Segoe UI", 10, "bold"), anchor="e").grid(
            row=row, column=1, sticky="ew", padx=12 if str(parent.cget("bg")) == "#F8FAFC" else 18, pady=5
        )

    def build_add_card(self, parent: tk.Widget, row: int, column: int) -> None:
        card = self.card(parent, row, column)
        card.grid_columnconfigure(0, weight=1)
        card.grid_rowconfigure(1, weight=1)

        header = tk.Frame(card, bg=CARD)
        header.grid(row=0, column=0, sticky="ew", padx=18, pady=(15, 6))
        header.grid_columnconfigure(0, weight=1)
        tk.Label(header, text="Add New Usernames", bg=CARD, fg=TEXT, font=("Segoe UI", 13, "bold")).grid(row=0, column=0, sticky="w")
        tk.Label(header, text="One per line. Comma or semicolon batches also work.", bg=CARD, fg=MUTED, font=("Segoe UI", 9)).grid(row=1, column=0, sticky="w", pady=(2, 0))

        self.add_box = tk.Text(card, height=4, wrap="word", bg="#F8FAFC", fg=TEXT, relief="solid", bd=1, font=("Segoe UI", 10), padx=10, pady=8)
        self.add_box.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 8))

        actions = tk.Frame(card, bg=CARD)
        actions.grid(row=2, column=0, sticky="ew", padx=18, pady=(0, 14))
        actions.grid_columnconfigure(0, weight=1)
        actions.grid_columnconfigure(1, weight=1)
        self.button(actions, "Save to Excel/CSV", ORANGE, ORANGE_HOVER, self.save_new_usernames).grid(row=0, column=0, sticky="ew", padx=(0, 6), ipady=6)
        self.button(actions, "Clear Box", SLATE, SLATE_HOVER, self.clear_add_box).grid(row=0, column=1, sticky="ew", padx=(6, 0), ipady=6)

    def button(self, parent: tk.Widget, text: str, bg: str, hover: str, command, width: int | None = None) -> tk.Button:
        button = tk.Button(
            parent,
            text=text,
            command=command,
            bg=bg,
            fg="white",
            activebackground=hover,
            activeforeground="white",
            relief="raised",
            bd=3,
            cursor="hand2",
            font=("Segoe UI", 10, "bold"),
            width=width or 0,
            highlightthickness=0,
        )
        button.bind("<Enter>", lambda event: button.configure(bg=hover))
        button.bind("<Leave>", lambda event: button.configure(bg=bg))
        return button

    def set_status(self, message: str, *, error: bool = False) -> None:
        self.status_var.set(message)
        if error:
            self.root.bell()

    def set_output(self, text: str) -> None:
        self.output_box.configure(state="normal")
        self.output_box.delete("1.0", tk.END)
        self.output_box.insert("1.0", text)
        self.output_box.configure(state="disabled")

    def update_metrics(self) -> None:
        self.total_var.set(str(len(self.usernames)))
        self.pick_var.set(str(len(self.result_text.split())) if self.result_text else "0")
        if self.loaded_file:
            self.file_var.set(os.path.basename(self.loaded_file))
        else:
            self.file_var.set("No sheet yet")

    def load_data_async(self, *, initial: bool) -> None:
        self.set_status("Loading usernames from the app folder...")
        self.file_var.set("Loading...")

        def worker() -> None:
            try:
                names, path, warning = load_usernames_from_folder(self.folder)
                error = None
            except Exception as exc:
                names, path, warning, error = [], None, None, str(exc)
            self.root.after(0, lambda: self.finish_load(names, path, warning, error, initial))

        threading.Thread(target=worker, daemon=True).start()

    def finish_load(self, names: list[str], path: str | None, warning: str | None, error: str | None, initial: bool) -> None:
        self.usernames = names
        self.loaded_file = path
        self.result_text = ""
        self.update_metrics()

        if error:
            self.set_output("")
            self.set_status(f"Load error: {error}", error=True)
        elif warning:
            self.set_output("")
            self.set_status(warning)
        else:
            self.set_status(f"Loaded {len(names)} usernames from {os.path.basename(path or '')}.")
            if self.count_var.get().strip():
                self.shuffle_names(automatic=True)
            else:
                self.set_output("")

        if initial:
            self.count_entry.focus_force()

    def on_count_changed(self, *_args) -> None:
        if not self._trace_ready:
            return
        self.shuffle_names(automatic=True)

    def shuffle_manual(self) -> None:
        self.shuffle_names(automatic=False)

    def shuffle_names(self, *, automatic: bool) -> None:
        raw_value = self.count_var.get().strip()

        if not raw_value:
            self.result_text = ""
            self.pick_var.set("0")
            self.set_output("")
            if self.usernames:
                self.set_status(f"Ready. {len(self.usernames)} usernames loaded.")
            return

        if not raw_value.isdigit():
            self.result_text = ""
            self.pick_var.set("0")
            self.set_output("")
            self.set_status("Enter a valid positive number.", error=not automatic)
            return

        amount = int(raw_value)
        if amount <= 0:
            self.result_text = ""
            self.pick_var.set("0")
            self.set_output("")
            self.set_status("Enter a number greater than 0.", error=not automatic)
            return

        if not self.usernames:
            self.result_text = ""
            self.pick_var.set("0")
            self.set_output("")
            self.set_status("No usernames loaded. Save usernames to create usernames.xlsx.", error=not automatic)
            return

        if amount > len(self.usernames):
            self.result_text = ""
            self.pick_var.set("0")
            self.set_output("")
            self.set_status("Pick count cannot be larger than loaded usernames.", error=not automatic)
            return

        selected = random.sample(self.usernames, amount)
        self.result_text = " ".join(selected)
        self.pick_var.set(str(amount))
        self.set_output(self.result_text)
        self.set_status(f"Generated {amount} shuffled username{'s' if amount != 1 else ''}.")

    def copy_result(self) -> None:
        if not self.result_text:
            self.set_status("Nothing to copy yet. Enter a number first.", error=True)
            return
        self.root.clipboard_clear()
        self.root.clipboard_append(self.result_text)
        self.root.update_idletasks()
        self.set_status("Result copied to clipboard.")
        self.count_entry.focus_force()

    def clear_add_box(self) -> None:
        self.add_box.delete("1.0", tk.END)
        self.set_status("Add box cleared.")

    def save_new_usernames(self) -> None:
        raw_text = self.add_box.get("1.0", tk.END)
        parsed = parse_new_usernames(raw_text)

        if not parsed:
            self.set_status("No new usernames to save.", error=True)
            return

        existing = {name.casefold() for name in self.usernames}
        seen_new: set[str] = set()
        unique_new: list[str] = []
        skipped = 0

        for name in parsed:
            key = name.casefold()
            if key in existing or key in seen_new:
                skipped += 1
                continue
            seen_new.add(key)
            unique_new.append(name)

        if not unique_new:
            self.set_status(f"No new usernames saved. {skipped} duplicate item(s) skipped.", error=True)
            return

        target_path = self.loaded_file
        if not target_path or target_path.lower().endswith(".xls"):
            target_path = os.path.join(self.folder, "usernames.xlsx")

        existing_snapshot = list(self.usernames)
        self.set_status("Saving new usernames...")

        def worker() -> None:
            try:
                saved_path = append_usernames_to_file(target_path, unique_new, existing_snapshot)
                error = None
            except Exception as exc:
                saved_path = target_path
                error = str(exc)
            self.root.after(0, lambda: self.finish_save(unique_new, skipped, saved_path, error))

        threading.Thread(target=worker, daemon=True).start()

    def finish_save(self, added: list[str], skipped: int, saved_path: str, error: str | None) -> None:
        if error:
            self.set_status(f"Save error: {error}", error=True)
            return

        self.loaded_file = saved_path
        self.usernames.extend(added)
        self.add_box.delete("1.0", tk.END)
        self.update_metrics()
        message = f"Saved {len(added)} new username{'s' if len(added) != 1 else ''} to {os.path.basename(saved_path)}."
        if skipped:
            message += f" Skipped {skipped} duplicate item(s)."
        self.set_status(message)
        if self.count_var.get().strip():
            self.shuffle_names(automatic=True)


def set_windows_app_id() -> None:
    if os.name != "nt":
        return
    try:
        import ctypes

        app_id = f"ShowravZaman.UsernameShuffler.{APP_VERSION}"
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(app_id)
    except Exception:
        pass


def main() -> None:
    set_windows_app_id()
    root = tk.Tk()
    UsernameShufflerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
